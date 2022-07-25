#!/usr/bin/python3
## -*- coding: utf-8 -*-

import json
import os
import argparse 
import openpyxl

parser = argparse.ArgumentParser(description='Перегоняет xslx в JSON для подготовки компендиума для The Witcher RPG')

parser.add_argument('--xlsx', 
    type=str, 
    help='Файл до эксельки', 
    dest = 'xlsx',
    required=True
)

args = parser.parse_args()

book = openpyxl.load_workbook(args.xlsx)
sheet = book.active

spells = []

for col in range(2, 999):
    
    if sheet.cell(row=col, column=1).value in ["", None]:
        break

    print(sheet.cell(row=col, column=1).value)

    description = sheet.cell(row=col, column=4).value
    description = description.replace("\n", "")
    description = description.replace('\xad', '')

    spells.append({
        "name": sheet.cell(row=col, column=1).value,
        "img": "",
        "type": "skills",
        "data": {
            "description": description,
            "stat": sheet.cell(row=col, column=2).value,
            "difficult": sheet.cell(row=col, column=3).value
        }
    })

cf = open('skills.json', 'w', encoding='utf-8')
cf.writelines(json.dumps(spells, ensure_ascii=False, indent=4, sort_keys=True))
cf.close()