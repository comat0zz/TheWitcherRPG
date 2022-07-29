#!/usr/bin/python3
## -*- coding: utf-8 -*-

import json
import os
import argparse 
import openpyxl

parser = argparse.ArgumentParser(description='Перегоняет xslx в JSON для подготовки компендиума для The Witcher RPG')

parser.add_argument('--xlsx', 
    type=str, 
    help='Файл до эксельки', 
    dest = 'xlsx',
    required=True
)

args = parser.parse_args()

book = openpyxl.load_workbook(args.xlsx)
sheet = book.active

spells = {}
spells['magicspell'] = []
spells['invocDruid'] = []
spells['invocPriest'] = []
spells['InvocHighPriest'] = []
spells['signs'] = []
spells['ritual'] = []
spells['injury'] = []


for col in range(2, 999):
    
    if sheet.cell(row=1, column=col).value in ["", None]:
        break

    print(sheet.cell(row=1, column=col).value)

    distantion_value = sheet.cell(row=11, column=col).value
    if distantion_value in ["", None]:
        distantion_value = 0

    costStamina = sheet.cell(row=6, column=col).value
    if costStamina in ["", None]:
        costStamina = 0

    category = sheet.cell(row=4, column=col).value
    if category == 'spells':
        category = 'magicspell'
    
    img = sheet.cell(row=5, column=col).value
    if img != "" and img is not None:
        img = "systems/TheWitcherRPG/assets%s" % img

    if category == 'magicspell':
        spells['magicspell'].append({
            "name": sheet.cell(row=1, column=col).value,
            "img": img,
            "type": "spell",
            "data": {
                "costStamina": costStamina,
                "source": "core",
                "effect": sheet.cell(row=3, column=col).value,
                "category": category,
                "components": [],
                "specData": {
                    "element": sheet.cell(row=8, column=col).value,
                    "LevelType": sheet.cell(row=9, column=col).value,
                    "DistantionType": sheet.cell(row=10, column=col).value,
                    "DistantionValue": distantion_value,
                    "defence": sheet.cell(row=12, column=col).value,
                    "duration": sheet.cell(row=13, column=col).value
                }
            }
        })
    
    if category == 'signs':
        spells[category].append({
            "name": sheet.cell(row=1, column=col).value,
            "img": img,
            "type": "spell",
            "data": {
                "costStamina": costStamina,
                "source": "core",
                "effect": sheet.cell(row=3, column=col).value,
                "category": category,
                "components": [],
                "specData": {
                    "element": sheet.cell(row=8, column=col).value,
                    "LevelType": sheet.cell(row=9, column=col).value,
                    "DistantionType": sheet.cell(row=10, column=col).value,
                    "DistantionValue": distantion_value,
                    "defence": sheet.cell(row=12, column=col).value,
                    "duration": sheet.cell(row=13, column=col).value
                }
            }
        })
    
    if category in ['invocPriest', 'invocDruid']:
        spells[category].append({
            "name": sheet.cell(row=1, column=col).value,
            "img": img,
            "type": "spell",
            "data": {
                "costStamina": costStamina,
                "source": "core",
                "effect": sheet.cell(row=3, column=col).value,
                "category": category,
                "components": [],
                "specData": {
                    #"element": sheet.cell(row=8, column=col).value,
                    "LevelType": sheet.cell(row=9, column=col).value,
                    "DistantionType": sheet.cell(row=10, column=col).value,
                    "DistantionValue": distantion_value,
                    "defence": sheet.cell(row=12, column=col).value,
                    "duration": sheet.cell(row=13, column=col).value
                }
            }
        })
    
    if category in ['ritual']:
        spells[category].append({
            "name": sheet.cell(row=1, column=col).value,
            "img": img,
            "type": "spell",
            "data": {
                "costStamina": costStamina,
                "source": "core",
                "effect": sheet.cell(row=3, column=col).value,
                "category": category,
                "components": [],
                "specData": {
                    #"element": sheet.cell(row=8, column=col).value,
                    "preparationTime" :  sheet.cell(row=9, column=col).value,
                    "LevelType": sheet.cell(row=12, column=col).value,
                    "forgeDC": sheet.cell(row=10, column=col).value,
                    #"DistantionType": sheet.cell(row=10, column=col).value,
                    #"DistantionValue": distantion_value,
                    #"defence": ,
                    "duration": sheet.cell(row=11, column=col).value
                }
            }
        })

    if category in ['injury']:
        spells[category].append({
            "name": sheet.cell(row=1, column=col).value,
            "img": img,
            "type": "spell",
            "data": {
                "costStamina": costStamina,
                "source": "core",
                "effect": sheet.cell(row=3, column=col).value,
                "category": category,
                "components": [],
                "specData": {
                    "hazard": sheet.cell(row=8, column=col).value,
                    "withdrawal": sheet.cell(row=9, column=col).value,
                    #"element": sheet.cell(row=8, column=col).value,
                    #"preparationTime" :  sheet.cell(row=9, column=col).value
                    #"LevelType": sheet.cell(row=12, column=col).value,
                    #"forgeDC": sheet.cell(row=10, column=col).value,
                    #"DistantionType": sheet.cell(row=10, column=col).value,
                    #"DistantionValue": distantion_value,
                    #"defence": ,
                    #"duration": sheet.cell(row=11, column=col).value
                }
            }
        })

    if category in ['InvocHighPriest']:
        spells[category].append({
            "name": sheet.cell(row=1, column=col).value,
            "img": img,
            "type": "spell",
            "data": {
                "costStamina": costStamina,
                "source": "core",
                "effect": sheet.cell(row=3, column=col).value,
                "category": category,
                "components": [],
                "specData": {
                    #"hazard": sheet.cell(row=8, column=col).value,
                    #"withdrawal": sheet.cell(row=9, column=col).value,
                    #"element": sheet.cell(row=8, column=col).value,
                    #"preparationTime" :  sheet.cell(row=9, column=col).value
                    "LevelType": sheet.cell(row=9, column=col).value,
                    #"forgeDC": sheet.cell(row=10, column=col).value,
                    #"DistantionType": sheet.cell(row=10, column=col).value,
                    #"DistantionValue": distantion_value,
                    #"defence": ,
                    "deities": "",
                    "duration": sheet.cell(row=13, column=col).value
                }
            }
        })
    

    

for key in spells:
    cf = open('tmp/%s.json' % key, 'w', encoding='utf-8')
    cf.writelines(json.dumps(spells[key], ensure_ascii=False, indent=4, sort_keys=True))
    cf.close()